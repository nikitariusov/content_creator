"""
СТРУКТУРА И АЛГОРИТМ

1. Получаем таблицу с характеристиками и сохраняем в словарь (ексель)
1.1 Выбор файла через проводник

2. Получаем шаблон с описанием
2.1 Ввод шаблона
2.2 Шаблон в формате txt через проводник

3. Сохраняем ключи из словаря с характеристиками

4. Копируем шаблон в новую переменную и циклом заменяем все характеристики в нем

СДЕЛАТЬ !!!
 1. Получение характеристик из таблицы
 2. Получение шаблона текста из файла тхт и ворд
 3. Сделать разветвление описаний
 4. Проверку по цифрам (гарантия 20 месяца) - найти библиотеку что это сделает
"""

import copy
import easygui  # Выбор файлов через проводник
import openpyxl
from openpyxl import Workbook

search_attr = []


def input_template():
    """Попытка, при запуске не работает, а в дебаге работает"""

    print('Вы можете ввести шалон описания вручную или выбрать файл с шаблоном в формате txt.\n'
          'Для выбора файла нажмите Enter или введите текст шаблона.\n'
          'Внимание! В шаблоне, все заменяемые занчения должны быть в фигурных кавычках {пример}...\n')

    user_input = str(input('... '))
    if user_input == '':
        template = get_template()


def get_template():
    file_path = easygui.fileopenbox('Выбери файл с шаблоном')
    f = open(file_path, 'r', encoding='utf-8')  # открываем файл для чтения
    template = f.read()  # читаем файл
    f.close()  # закрываем работу с файлом!!!
    return template


def create_content(attrib_list, template):
    """ Копируем шаблон описания в новую переменную и заменяем в нем
    ключи (наименования) характеристик на их начения из словаря с характеристиками """
    data = []

    for char in attrib_list:
        content = copy.copy(template)
        #   print('\n[TEST] char', char)
        try:
            for key in char.keys():
                value = str(char[key])
                #   print(value)
                content = content.replace('{' + key + '}', value)
        except KeyError:
            print(f'Ошибка, не может найти элемент: {char}')
        print('\n', content)
        data.append(content)
    return data


def read_excel(file_path):
    file_xl = openpyxl.load_workbook(file_path)
    ws = file_xl.active
    max_column = ws.max_column
    print(f'\nКоличество товаров: {max_column}.')

    #   for row in ws.values:
    #   print(row)

    attr_key = []
    for i in range(1, ws.max_column + 1):
        cols_item = ws.cell(row=1, column=i)
        print(f'\tХарактеристика: {cols_item.value}')

        attr_key.append(cols_item.value)

    all_data = []
    for k in range(2, ws.max_row + 1):
        attr = {}
        n = 0
        for j in range(1, ws.max_column + 1):
            cols_value = ws.cell(row=k, column=j)
            #   print('пробую вывести колонки:', cols_value.value)
            attr[attr_key[n]] = cols_value.value
            n += 1
        all_data.append(attr)

    #   print(all_data)
    return all_data


def rec_to_file(contents):
    wb = Workbook()
    file_name = "Data.xlsx"
    ws1 = wb.active
    ws1.title = "Data_base"

    ws1['A1'] = 'Описание'

    row = 2
    for content in contents:

        ws1[f'A{row}'] = content
        row += 1

    wb.save(filename=file_name)
    print(f'\nОбработка завершена. Сохранен файл {file_name}.')


def main():
    """ Главная функция программы """

    template = get_template()
    print('\n[INFO] Шаблон с описанием:')
    print('\t', template)
    file_path = easygui.fileopenbox('Выбери файл с характеристиками')
    data_from_excel = read_excel(file_path)
    data = create_content(data_from_excel, template)
    rec_to_file(data)


main()
